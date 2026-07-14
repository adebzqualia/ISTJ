"""Low-level helpers shared across the extractor."""

from __future__ import annotations

import hashlib
import json
import math
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl.utils import get_column_letter, range_boundaries


WHITESPACE_PATTERN = re.compile(r"\s+")
NON_WORD_PATTERN = re.compile(r"[^a-z0-9]+")


def file_sha256(path: Path, chunk_size: int = 1024 * 1024) -> str:
    """Return a SHA-256 digest for a file."""

    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(chunk_size), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalize_label(value: Any) -> str:
    """Normalize a user-facing label for matching."""

    if value is None:
        return ""
    text = WHITESPACE_PATTERN.sub(" ", str(value).strip())
    return text.casefold()


def technical_name(value: Any, fallback: str) -> str:
    """Create a stable snake-like technical name."""

    normalized = normalize_label(value)
    name = NON_WORD_PATTERN.sub("_", normalized).strip("_")
    return name or fallback


def json_dumps(value: Any) -> str:
    """Serialize a value consistently for tabular output."""

    return json.dumps(value, ensure_ascii=False, default=str)


def is_missing(value: Any) -> bool:
    """Return whether a value should be treated as empty."""

    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def is_number(value: Any) -> bool:
    """Return whether a value is a non-boolean number."""

    return isinstance(value, (int, float)) and not isinstance(value, bool)


def is_date_like(value: Any) -> bool:
    """Return whether a value is a Python date or datetime."""

    return isinstance(value, (date, datetime))


def a1_range(
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> str:
    """Return an A1 range from numeric bounds."""

    return (
        f"{get_column_letter(min_col)}{min_row}:"
        f"{get_column_letter(max_col)}{max_row}"
    )


def parse_a1_range(value: str) -> tuple[int, int, int, int]:
    """Return row and column bounds from an A1 range."""

    min_col, min_row, max_col, max_row = range_boundaries(value)
    return min_row, max_row, min_col, max_col


def flatten(values: Iterable[Iterable[Any]]) -> list[Any]:
    """Flatten a two-dimensional iterable."""

    return [item for row in values for item in row]
