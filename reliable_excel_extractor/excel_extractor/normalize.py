"""Normalization of accepted candidate blocks."""

from __future__ import annotations

from typing import Any

from openpyxl.utils import get_column_letter

from .config import ExtractionConfig
from .interpret import effective_value
from .models import BlockInterpretation, ExtractionWarning, WarningSeverity
from .periods import parse_period_label
from .utils import is_missing, is_number, json_dumps, normalize_label, technical_name


def _lookup(
    raw_cells: list[dict[str, Any]],
) -> dict[tuple[int, int], dict[str, Any]]:
    """Index raw cells by numeric coordinate."""

    return {
        (record["row_index"], record["column_index"]): record
        for record in raw_cells
    }




def _measure_value(record: dict[str, Any] | None) -> Any:
    """Return a value safe for analytical measures.

    Formula text is never used as a value. A formula contributes only when a
    cached result is available.
    """

    if not record:
        return None
    if record.get("formula") is not None:
        return record.get("cached_value")
    return record.get("raw_value")


def _canonical_kpi(value: Any, aliases: dict[str, str]) -> Any:
    """Map a KPI alias to a canonical label when configured."""

    normalized = normalize_label(value)
    return aliases.get(normalized, value)


def _header_sources(
    interpretation: BlockInterpretation,
    column: int,
) -> list[str]:
    """Return source addresses used for one header path."""

    if interpretation.header_depth == 0:
        return []
    return [
        f"{get_column_letter(column)}{row}"
        for row in range(
            interpretation.header_start_row,
            interpretation.header_end_row + 1,
        )
    ]


def _resolved_row_labels(
    interpretation: BlockInterpretation,
    lookup: dict[tuple[int, int], dict[str, Any]],
    row: int,
    prior_labels: list[Any],
    config: ExtractionConfig,
) -> tuple[list[Any], list[str], list[str]]:
    """Resolve hierarchical row labels while retaining methods and sources."""

    labels: list[Any] = []
    methods: list[str] = []
    sources: list[str] = []
    for offset in range(interpretation.row_label_columns):
        column = interpretation.candidate.min_col + offset
        record = lookup.get((row, column))
        value = effective_value(record)
        source = f"{get_column_letter(column)}{row}"
        method = "original"
        if record and record.get("is_merged") and not is_missing(
            record.get("resolved_merged_value")
        ):
            method = "merged_range"
        elif (
            is_missing(value)
            and config.propagate_blank_hierarchy_labels
            and offset < len(prior_labels)
            and not is_missing(prior_labels[offset])
        ):
            value = prior_labels[offset]
            method = "heuristic_previous_row"
        labels.append(value)
        methods.append(method)
        sources.append(source)
    return labels, methods, sources


def _observation_record(
    *,
    interpretation: BlockInterpretation,
    source_record: dict[str, Any],
    value: Any,
    kpi_name: Any,
    period: dict[str, Any],
    scenario: Any,
    unit: Any,
    currency: Any,
    dimensions: dict[str, Any],
    row_labels: list[Any],
    row_label_sources: list[str],
    row_label_methods: list[str],
    header_path: list[Any],
    header_sources: list[str],
    workbook_id: str,
) -> dict[str, Any]:
    """Build one fully traced analytical observation."""

    row_type = interpretation.row_types.get(
        source_record["row_index"], "unknown"
    )
    return {
        "workbook_id": workbook_id,
        "observation_id": (
            f"{interpretation.candidate.block_id}__"
            f"{source_record['cell_address']}"
        ),
        "block_id": interpretation.candidate.block_id,
        "sheet_name": interpretation.candidate.sheet_name,
        "kpi_name": kpi_name,
        "period_label_original": period.get("period_label_original"),
        "period_type": period.get("period_type"),
        "period_start": period.get("period_start"),
        "period_end": period.get("period_end"),
        "period_year": period.get("period_year"),
        "period_number": period.get("period_number"),
        "period_parse_status": period.get("period_parse_status"),
        "period_parse_confidence": period.get("period_parse_confidence"),
        "scenario": scenario or period.get("scenario"),
        "unit": unit,
        "currency": currency,
        "value": value,
        "value_type": source_record.get("semantic_type"),
        "is_formula": source_record.get("formula") is not None,
        "is_total": row_type in {"subtotal", "total", "grand_total"},
        "row_type": row_type,
        "dimensions_json": json_dumps(dimensions),
        "row_label_path_original_json": json_dumps(row_labels),
        "header_path_original_json": json_dumps(header_path),
        "source_sheet": source_record["sheet_name"],
        "source_range": interpretation.candidate.source_range,
        "source_value_cell": source_record["cell_address"],
        "source_value_formula": source_record.get("formula"),
        "source_header_cells_json": json_dumps(header_sources),
        "source_row_label_cells_json": json_dumps(row_label_sources),
        "row_label_resolution_methods_json": json_dumps(row_label_methods),
        "extraction_confidence": interpretation.confidence,
        "warning_codes_json": source_record.get("warning_codes_json", "[]"),
    }


def _normalize_cross_tab(
    interpretation: BlockInterpretation,
    raw_cells: list[dict[str, Any]],
    config: ExtractionConfig,
    workbook_id: str,
) -> tuple[list[dict[str, Any]], list[ExtractionWarning]]:
    """Normalize a cross-tab matrix into long-form observations."""

    lookup = _lookup(raw_cells)
    observations: list[dict[str, Any]] = []
    warnings: list[ExtractionWarning] = []
    prior_labels: list[Any] = []
    aliases = config.template.known_kpi_aliases

    first_measure_col = (
        interpretation.candidate.min_col + interpretation.row_label_columns
    )
    for row in range(
        interpretation.data_start_row,
        interpretation.candidate.max_row + 1,
    ):
        labels, methods, label_sources = _resolved_row_labels(
            interpretation,
            lookup,
            row,
            prior_labels,
            config,
        )
        if any(not is_missing(value) for value in labels):
            prior_labels = labels
        non_missing_label_indexes = [
            index for index, value in enumerate(labels) if not is_missing(value)
        ]
        kpi_index = non_missing_label_indexes[-1] if non_missing_label_indexes else None
        kpi_name = labels[kpi_index] if kpi_index is not None else None
        dimensions = {
            f"dimension_{index + 1}": value
            for index, value in enumerate(labels)
            if kpi_index is not None
            and index < kpi_index
            and not is_missing(value)
        }
        kpi_name = _canonical_kpi(kpi_name, aliases)

        for column in range(
            first_measure_col,
            interpretation.candidate.max_col + 1,
        ):
            source = lookup.get((row, column))
            if not source:
                continue
            value = _measure_value(source)
            if is_missing(value):
                continue
            if not is_number(value):
                warnings.append(
                    ExtractionWarning(
                        warning_code="BL004",
                        severity=WarningSeverity.WARNING,
                        message="Non-numeric value found in a matrix measure cell.",
                        workbook_id=workbook_id,
                        sheet_name=interpretation.candidate.sheet_name,
                        block_id=interpretation.candidate.block_id,
                        cell_address=source["cell_address"],
                        source_range=interpretation.candidate.source_range,
                        suggested_action="Review the measure-column interpretation.",
                    )
                )
                continue

            header_path = interpretation.header_paths.get(column, [])
            period_candidates = [
                parse_period_label(item)
                for item in reversed(header_path)
            ]
            period = next(
                (
                    item
                    for item in period_candidates
                    if item["period_parse_confidence"] >= 0.75
                    and item["period_type"] is not None
                ),
                parse_period_label(
                    header_path[-1] if header_path else None
                ),
            )
            scenario = next(
                (
                    item["scenario"]
                    for item in period_candidates
                    if item["scenario"]
                ),
                None,
            )
            observations.append(
                _observation_record(
                    interpretation=interpretation,
                    source_record=source,
                    value=value,
                    kpi_name=kpi_name,
                    period=period,
                    scenario=scenario,
                    unit=None,
                    currency=None,
                    dimensions=dimensions,
                    row_labels=labels,
                    row_label_sources=label_sources,
                    row_label_methods=methods,
                    header_path=header_path,
                    header_sources=_header_sources(interpretation, column),
                    workbook_id=workbook_id,
                )
            )
    return observations, warnings


def _normalize_record_table(
    interpretation: BlockInterpretation,
    raw_cells: list[dict[str, Any]],
    config: ExtractionConfig,
    workbook_id: str,
) -> tuple[list[dict[str, Any]], list[ExtractionWarning]]:
    """Normalize a record-oriented table into measure observations."""

    lookup = _lookup(raw_cells)
    observations: list[dict[str, Any]] = []
    aliases = config.template.known_kpi_aliases

    technical_names: dict[int, str] = {}
    seen: dict[str, int] = {}
    for column, path in interpretation.header_paths.items():
        label = path[-1] if path else f"column_{column}"
        base = technical_name(label, f"column_{column}")
        seen[base] = seen.get(base, 0) + 1
        technical_names[column] = (
            base if seen[base] == 1 else f"{base}_{seen[base]}"
        )

    for row in range(
        interpretation.data_start_row,
        interpretation.candidate.max_row + 1,
    ):
        row_values = {
            technical_names[column]: effective_value(lookup.get((row, column)))
            for column in range(
                interpretation.candidate.min_col,
                interpretation.candidate.max_col + 1,
            )
        }
        role_values: dict[str, Any] = {}
        dimensions: dict[str, Any] = {}
        for column, role in interpretation.column_roles.items():
            name = technical_names[column]
            value = row_values[name]
            if role in {"kpi", "period", "scenario", "unit", "currency"}:
                role_values[role] = value
            elif role == "dimension":
                dimensions[name] = value

        kpi_name = _canonical_kpi(role_values.get("kpi"), aliases)
        period = parse_period_label(role_values.get("period"))
        for column, role in interpretation.column_roles.items():
            if role != "measure":
                continue
            source = lookup.get((row, column))
            if not source:
                continue
            value = _measure_value(source)
            if is_missing(value) or not is_number(value):
                continue
            measure_header = interpretation.header_paths.get(column, [])
            measure_name = measure_header[-1] if measure_header else None
            effective_kpi = kpi_name or _canonical_kpi(measure_name, aliases)
            measure_period = period
            if measure_period.get("period_parse_confidence", 0.0) < 0.75:
                measure_period = next(
                    (
                        parsed
                        for parsed in [
                            parse_period_label(item)
                            for item in reversed(measure_header)
                        ]
                        if parsed["period_parse_confidence"] >= 0.75
                        and parsed["period_type"] is not None
                    ),
                    measure_period,
                )
            observations.append(
                _observation_record(
                    interpretation=interpretation,
                    source_record=source,
                    value=value,
                    kpi_name=effective_kpi,
                    period=measure_period,
                    scenario=role_values.get("scenario"),
                    unit=role_values.get("unit"),
                    currency=role_values.get("currency"),
                    dimensions=dimensions,
                    row_labels=list(dimensions.values()),
                    row_label_sources=[],
                    row_label_methods=[],
                    header_path=measure_header,
                    header_sources=_header_sources(interpretation, column),
                    workbook_id=workbook_id,
                )
            )
    return observations, []


def _normalize_kpi_value_block(
    interpretation: BlockInterpretation,
    raw_cells: list[dict[str, Any]],
    config: ExtractionConfig,
    workbook_id: str,
) -> tuple[list[dict[str, Any]], list[ExtractionWarning]]:
    """Normalize a compact KPI/value block."""

    lookup = _lookup(raw_cells)
    observations: list[dict[str, Any]] = []
    aliases = config.template.known_kpi_aliases
    label_column = interpretation.candidate.min_col
    measure_columns = [
        column
        for column, role in interpretation.column_roles.items()
        if role == "measure"
    ]
    if not measure_columns and interpretation.candidate.width >= 2:
        measure_columns = [label_column + 1]

    for row in range(
        interpretation.data_start_row,
        interpretation.candidate.max_row + 1,
    ):
        label_record = lookup.get((row, label_column))
        kpi_name = _canonical_kpi(effective_value(label_record), aliases)
        if is_missing(kpi_name):
            continue
        for column in measure_columns:
            source = lookup.get((row, column))
            if not source:
                continue
            value = _measure_value(source)
            if is_missing(value) or not is_number(value):
                continue
            header_path = interpretation.header_paths.get(column, [])
            observations.append(
                _observation_record(
                    interpretation=interpretation,
                    source_record=source,
                    value=value,
                    kpi_name=kpi_name,
                    period=parse_period_label(
                        header_path[-1] if header_path else None
                    ),
                    scenario=None,
                    unit=None,
                    currency=None,
                    dimensions={},
                    row_labels=[kpi_name],
                    row_label_sources=[
                        label_record["cell_address"] if label_record else None
                    ],
                    row_label_methods=["original"],
                    header_path=header_path,
                    header_sources=_header_sources(interpretation, column),
                    workbook_id=workbook_id,
                )
            )
    return observations, []


def normalize_block(
    interpretation: BlockInterpretation,
    raw_cells: list[dict[str, Any]],
    config: ExtractionConfig,
    workbook_id: str,
) -> tuple[list[dict[str, Any]], list[ExtractionWarning]]:
    """Normalize one accepted block according to its structural class."""

    if interpretation.status != "accepted":
        return [], []
    if interpretation.block_class == "cross_tab_matrix":
        return _normalize_cross_tab(
            interpretation, raw_cells, config, workbook_id
        )
    if interpretation.block_class == "rectangular_record_table":
        return _normalize_record_table(
            interpretation, raw_cells, config, workbook_id
        )
    if interpretation.block_class == "kpi_value_block":
        return _normalize_kpi_value_block(
            interpretation, raw_cells, config, workbook_id
        )
    return [], []


def block_column_records(
    interpretation: BlockInterpretation,
    workbook_id: str,
) -> list[dict[str, Any]]:
    """Build one metadata row per interpreted block column."""

    records: list[dict[str, Any]] = []
    seen_names: dict[str, int] = {}
    for column in range(
        interpretation.candidate.min_col,
        interpretation.candidate.max_col + 1,
    ):
        path = interpretation.header_paths.get(column, [])
        base = technical_name(
            "__".join(str(value) for value in path if not is_missing(value)),
            f"column_{column}",
        )
        seen_names[base] = seen_names.get(base, 0) + 1
        name = base if seen_names[base] == 1 else f"{base}_{seen_names[base]}"
        period = next(
            (
                parsed
                for parsed in [
                    parse_period_label(value) for value in reversed(path)
                ]
                if parsed["period_parse_confidence"] >= 0.75
            ),
            parse_period_label(None),
        )
        records.append(
            {
                "workbook_id": workbook_id,
                "block_id": interpretation.candidate.block_id,
                "sheet_name": interpretation.candidate.sheet_name,
                "column_index": column,
                "source_column_letter": get_column_letter(column),
                "header_path_original_json": json_dumps(path),
                "technical_name": name,
                "semantic_role": interpretation.column_roles.get(
                    column, "unknown"
                ),
                "period_metadata_json": json_dumps(period),
                "column_confidence": interpretation.confidence,
            }
        )
    return records


def block_row_records(
    interpretation: BlockInterpretation,
    raw_cells: list[dict[str, Any]],
    config: ExtractionConfig,
    workbook_id: str,
) -> list[dict[str, Any]]:
    """Build one metadata row per interpreted body row."""

    lookup = _lookup(raw_cells)
    prior_labels: list[Any] = []
    records: list[dict[str, Any]] = []
    for row in range(
        interpretation.data_start_row,
        interpretation.candidate.max_row + 1,
    ):
        labels, methods, sources = _resolved_row_labels(
            interpretation,
            lookup,
            row,
            prior_labels,
            config,
        )
        if any(not is_missing(value) for value in labels):
            prior_labels = labels
        records.append(
            {
                "workbook_id": workbook_id,
                "block_id": interpretation.candidate.block_id,
                "sheet_name": interpretation.candidate.sheet_name,
                "row_index": row,
                "row_label_path_original_json": json_dumps(labels),
                "row_label_source_cells_json": json_dumps(sources),
                "row_label_resolution_methods_json": json_dumps(methods),
                "row_type": interpretation.row_types.get(row, "unknown"),
                "row_confidence": interpretation.confidence,
            }
        )
    return records
