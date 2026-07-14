"""Workbook loading, inventory, and raw-cell extraction."""

from __future__ import annotations

import re
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .config import ExtractionConfig
from .models import ExtractionWarning, WarningSeverity
from .utils import (
    a1_range,
    file_sha256,
    is_date_like,
    is_missing,
    is_number,
    json_dumps,
    normalize_label,
)


FORMULA_REF_PATTERN = re.compile(
    r"(?:(?:'(?P<quoted>[^']+)'|(?P<plain>[A-Za-z0-9_ .-]+))!)?"
    r"(?P<range>\$?[A-Z]{1,3}\$?\d+(?::\$?[A-Z]{1,3}\$?\d+)?)"
)


def load_workbook_pair(path: Path) -> tuple[Workbook, Workbook]:
    """Load formula and cached-value views of an Excel workbook.

    :param path: Path to an ``.xlsx`` or ``.xlsm`` workbook.
    :return: Formula and cached-value workbook pair.
    :raises ValueError: If the extension is unsupported.
    """

    suffix = path.suffix.casefold()
    if suffix not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        raise ValueError(
            "Only .xlsx, .xlsm, .xltx, and .xltm workbooks are supported."
        )

    keep_vba = suffix in {".xlsm", ".xltm"}
    formula_workbook = load_workbook(
        path,
        data_only=False,
        read_only=False,
        keep_vba=keep_vba,
        keep_links=True,
    )
    value_workbook = load_workbook(
        path,
        data_only=True,
        read_only=False,
        keep_vba=keep_vba,
        keep_links=True,
    )
    return formula_workbook, value_workbook


def workbook_inventory(
    path: Path,
    workbook: Workbook,
    workbook_id: str,
) -> dict[str, Any]:
    """Build workbook-level inventory metadata."""

    calculation = getattr(workbook, "calculation", None)
    defined_names = [item.name for item in workbook.defined_names.values()]
    return {
        "workbook_id": workbook_id,
        "file_name": path.name,
        "file_path": str(path.resolve()),
        "file_size_bytes": path.stat().st_size,
        "file_sha256": file_sha256(path),
        "sheet_count": len(workbook.sheetnames),
        "sheet_names_json": json_dumps(workbook.sheetnames),
        "has_macros": path.suffix.casefold() in {".xlsm", ".xltm"},
        "has_external_links": bool(getattr(workbook, "_external_links", [])),
        "calculation_mode": getattr(calculation, "calcMode", None),
        "full_calculation_on_load": getattr(
            calculation, "fullCalcOnLoad", None
        ),
        "force_full_calculation": getattr(calculation, "forceFullCalc", None),
        "defined_names_json": json_dumps(sorted(defined_names)),
        "extraction_timestamp_utc": datetime.now(timezone.utc).isoformat(),
    }


def build_merged_cell_map(worksheet: Worksheet) -> dict[str, dict[str, Any]]:
    """Map every cell in a merged range to its anchor and value."""

    merged_map: dict[str, dict[str, Any]] = {}
    for merged_range in worksheet.merged_cells.ranges:
        anchor = worksheet.cell(
            row=merged_range.min_row,
            column=merged_range.min_col,
        )
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for column in range(
                merged_range.min_col,
                merged_range.max_col + 1,
            ):
                address = f"{get_column_letter(column)}{row}"
                merged_map[address] = {
                    "merged_range": str(merged_range),
                    "merged_anchor": anchor.coordinate,
                    "resolved_merged_value": anchor.value,
                }
    return merged_map


def _cell_has_relevant_content(
    cell: Any,
    *,
    include_styled_empty_cells: bool,
) -> bool:
    """Return whether a cell belongs in the raw-cell output."""

    if isinstance(cell, MergedCell):
        return False
    if not is_missing(cell.value):
        return True
    if cell.comment is not None or cell.hyperlink is not None:
        return True
    return include_styled_empty_cells and cell.has_style


def actual_bounds(
    worksheet: Worksheet,
    merged_map: dict[str, dict[str, Any]],
    *,
    include_styled_empty_cells: bool,
) -> tuple[int | None, int | None, int | None, int | None]:
    """Compute the real occupied worksheet bounds."""

    coordinates: list[tuple[int, int]] = []
    for cell in worksheet._cells.values():  # noqa: SLF001 - intentional scan
        if _cell_has_relevant_content(
            cell,
            include_styled_empty_cells=include_styled_empty_cells,
        ):
            coordinates.append((cell.row, cell.column))

    for address in merged_map:
        cell = worksheet[address]
        coordinates.append((cell.row, cell.column))

    if not coordinates:
        return None, None, None, None

    rows, columns = zip(*coordinates)
    return min(rows), max(rows), min(columns), max(columns)


def _color_record(color: Any) -> dict[str, Any] | None:
    """Return a serializable OpenPyXL color representation."""

    if color is None:
        return None
    return {
        "type": getattr(color, "type", None),
        "rgb": getattr(color, "rgb", None),
        "indexed": getattr(color, "indexed", None),
        "theme": getattr(color, "theme", None),
        "tint": getattr(color, "tint", None),
    }


def _style_record(cell: Any) -> dict[str, Any]:
    """Return simplified style metadata for one cell."""

    return {
        "font": {
            "name": cell.font.name,
            "size": cell.font.sz,
            "bold": bool(cell.font.bold),
            "italic": bool(cell.font.italic),
            "underline": cell.font.underline,
            "color": _color_record(cell.font.color),
        },
        "fill": {
            "fill_type": cell.fill.fill_type,
            "foreground": _color_record(cell.fill.fgColor),
            "background": _color_record(cell.fill.bgColor),
        },
        "alignment": {
            "horizontal": cell.alignment.horizontal,
            "vertical": cell.alignment.vertical,
            "indent": cell.alignment.indent,
            "wrap_text": cell.alignment.wrap_text,
        },
        "border": {
            "left": cell.border.left.style,
            "right": cell.border.right.style,
            "top": cell.border.top.style,
            "bottom": cell.border.bottom.style,
        },
    }


def semantic_type(
    raw_value: Any,
    cached_value: Any,
    data_type: str,
) -> str:
    """Classify a cell into a simple semantic type."""

    if data_type == "f":
        target = cached_value
        if is_missing(target):
            return "FORMULA_UNKNOWN"
        if is_number(target):
            return "FORMULA_NUMBER"
        if is_date_like(target):
            return "FORMULA_DATE"
        if isinstance(target, str):
            return "FORMULA_TEXT"
        return "FORMULA_OTHER"

    value = raw_value
    if is_missing(value):
        return "EMPTY"
    if data_type == "e":
        return "ERROR"
    if isinstance(value, bool):
        return "BOOLEAN"
    if is_date_like(value):
        return "DATE"
    if is_number(value):
        return "NUMBER"
    if isinstance(value, str):
        return "TEXT"
    return "OTHER"


def extract_formula_references(
    formula: str | None,
    current_sheet: str,
) -> list[dict[str, str]]:
    """Extract basic A1 references from a formula string.

    The parser intentionally records only direct A1 references. Structured
    references, dynamic arrays, and indirect references remain in the formula
    text and should be treated as unresolved dependencies.
    """

    if not formula or not formula.startswith("="):
        return []

    references: list[dict[str, str]] = []
    for match in FORMULA_REF_PATTERN.finditer(formula):
        sheet = match.group("quoted") or match.group("plain") or current_sheet
        references.append(
            {
                "sheet": sheet.strip(),
                "range": match.group("range").replace("$", ""),
            }
        )
    return references


def extract_raw_cells(
    formula_sheet: Worksheet,
    value_sheet: Worksheet,
    workbook_id: str,
    config: ExtractionConfig,
) -> tuple[list[dict[str, Any]], list[ExtractionWarning]]:
    """Extract source-faithful records for relevant worksheet cells."""

    merged_map = build_merged_cell_map(formula_sheet)
    min_row, max_row, min_col, max_col = actual_bounds(
        formula_sheet,
        merged_map,
        include_styled_empty_cells=config.include_styled_empty_cells,
    )
    if min_row is None:
        return [], []

    records: list[dict[str, Any]] = []
    warnings: list[ExtractionWarning] = []

    relevant_addresses = set(merged_map)
    for cell in formula_sheet._cells.values():  # noqa: SLF001
        if _cell_has_relevant_content(
            cell,
            include_styled_empty_cells=config.include_styled_empty_cells,
        ):
            relevant_addresses.add(cell.coordinate)

    for address in sorted(
        relevant_addresses,
        key=lambda item: (
            formula_sheet[item].row,
            formula_sheet[item].column,
        ),
    ):
        formula_cell = formula_sheet[address]
        value_cell = value_sheet[address]
        merge_info = merged_map.get(address, {})
        raw_value = (
            None if isinstance(formula_cell, MergedCell) else formula_cell.value
        )
        cached_value = (
            None if isinstance(value_cell, MergedCell) else value_cell.value
        )
        formula = raw_value if formula_cell.data_type == "f" else None
        is_error = formula_cell.data_type == "e" or value_cell.data_type == "e"
        warning_codes: list[str] = []

        if formula and is_missing(cached_value):
            warning_codes.append("CL001")
            warnings.append(
                ExtractionWarning(
                    warning_code="CL001",
                    severity=WarningSeverity.WARNING,
                    message="Formula cell has no cached value.",
                    workbook_id=workbook_id,
                    sheet_name=formula_sheet.title,
                    cell_address=address,
                    suggested_action=(
                        "Recalculate and save the workbook in Excel or a "
                        "compatible calculation engine."
                    ),
                )
            )
        if is_error:
            warning_codes.append("CL002")
            warnings.append(
                ExtractionWarning(
                    warning_code="CL002",
                    severity=WarningSeverity.ERROR,
                    message="Excel error value detected.",
                    workbook_id=workbook_id,
                    sheet_name=formula_sheet.title,
                    cell_address=address,
                    suggested_action="Review the source formula or input cell.",
                )
            )

        style = (
            _style_record(formula_cell)
            if config.include_style_details
            and not isinstance(formula_cell, MergedCell)
            else {}
        )
        row_dimension = formula_sheet.row_dimensions[formula_cell.row]
        column_letter = get_column_letter(formula_cell.column)
        column_dimension = formula_sheet.column_dimensions[column_letter]

        records.append(
            {
                "workbook_id": workbook_id,
                "sheet_name": formula_sheet.title,
                "row_index": formula_cell.row,
                "column_index": formula_cell.column,
                "column_letter": column_letter,
                "cell_address": address,
                "raw_value": raw_value,
                "cached_value": cached_value,
                "formula": formula,
                "formula_references_json": json_dumps(
                    extract_formula_references(formula, formula_sheet.title)
                ),
                "data_type": formula_cell.data_type,
                "semantic_type": semantic_type(
                    raw_value, cached_value, formula_cell.data_type
                ),
                "number_format": (
                    None
                    if isinstance(formula_cell, MergedCell)
                    else formula_cell.number_format
                ),
                "is_merged": address in merged_map,
                "merged_range": merge_info.get("merged_range"),
                "merged_anchor": merge_info.get("merged_anchor"),
                "resolved_merged_value": merge_info.get(
                    "resolved_merged_value"
                ),
                "is_hidden_row": bool(row_dimension.hidden),
                "is_hidden_column": bool(column_dimension.hidden),
                "has_comment": (
                    False
                    if isinstance(formula_cell, MergedCell)
                    else formula_cell.comment is not None
                ),
                "comment_text": (
                    formula_cell.comment.text
                    if config.include_comments
                    and not isinstance(formula_cell, MergedCell)
                    and formula_cell.comment is not None
                    else None
                ),
                "has_hyperlink": (
                    False
                    if isinstance(formula_cell, MergedCell)
                    else formula_cell.hyperlink is not None
                ),
                "hyperlink_target": (
                    getattr(formula_cell.hyperlink, "target", None)
                    if not isinstance(formula_cell, MergedCell)
                    else None
                ),
                "style_id": (
                    None
                    if isinstance(formula_cell, MergedCell)
                    else formula_cell.style_id
                ),
                "is_bold": style.get("font", {}).get("bold", False),
                "horizontal_alignment": style.get("alignment", {}).get(
                    "horizontal"
                ),
                "vertical_alignment": style.get("alignment", {}).get(
                    "vertical"
                ),
                "fill_json": json_dumps(style.get("fill", {})),
                "font_json": json_dumps(style.get("font", {})),
                "border_json": json_dumps(style.get("border", {})),
                "warning_codes_json": json_dumps(warning_codes),
            }
        )

    return records, warnings


def inspect_sheet(
    worksheet: Worksheet,
    raw_cells: list[dict[str, Any]],
    workbook_id: str,
) -> dict[str, Any]:
    """Build worksheet-level inventory metadata."""

    if raw_cells:
        rows = [record["row_index"] for record in raw_cells]
        columns = [record["column_index"] for record in raw_cells]
        bounds = (min(rows), max(rows), min(columns), max(columns))
    else:
        bounds = (None, None, None, None)

    hidden_rows = [
        index
        for index, dimension in worksheet.row_dimensions.items()
        if dimension.hidden
    ]
    hidden_columns = [
        key
        for key, dimension in worksheet.column_dimensions.items()
        if dimension.hidden
    ]
    table_records = [
        {"name": name, "range": worksheet.tables[name].ref}
        for name in worksheet.tables
    ]

    return {
        "workbook_id": workbook_id,
        "sheet_name": worksheet.title,
        "sheet_index": worksheet.parent.sheetnames.index(worksheet.title),
        "visibility": worksheet.sheet_state,
        "actual_min_row": bounds[0],
        "actual_max_row": bounds[1],
        "actual_min_column": bounds[2],
        "actual_max_column": bounds[3],
        "actual_range": (
            a1_range(bounds[0], bounds[1], bounds[2], bounds[3])
            if all(value is not None for value in bounds)
            else None
        ),
        "reported_max_row": worksheet.max_row,
        "reported_max_column": worksheet.max_column,
        "non_empty_cell_count": sum(
            1
            for record in raw_cells
            if not is_missing(record["raw_value"])
            or record["formula"] is not None
            or not is_missing(record["resolved_merged_value"])
        ),
        "raw_cell_record_count": len(raw_cells),
        "formula_cell_count": sum(
            record["formula"] is not None for record in raw_cells
        ),
        "merged_range_count": len(worksheet.merged_cells.ranges),
        "merged_ranges_json": json_dumps(
            [str(item) for item in worksheet.merged_cells.ranges]
        ),
        "hidden_row_count": len(hidden_rows),
        "hidden_rows_json": json_dumps(hidden_rows),
        "hidden_column_count": len(hidden_columns),
        "hidden_columns_json": json_dumps(hidden_columns),
        "official_table_count": len(table_records),
        "official_tables_json": json_dumps(table_records),
        "freeze_panes": str(worksheet.freeze_panes or ""),
        "auto_filter_range": worksheet.auto_filter.ref,
        "comment_count": sum(record["has_comment"] for record in raw_cells),
        "error_cell_count": sum(
            record["semantic_type"] == "ERROR" for record in raw_cells
        ),
    }


def anchor_warnings(
    worksheet: Worksheet,
    raw_cells: list[dict[str, Any]],
    workbook_id: str,
    expected_labels: tuple[str, ...],
) -> list[ExtractionWarning]:
    """Return warnings for expected labels absent from a sheet."""

    available = {
        normalize_label(
            record["resolved_merged_value"]
            if not is_missing(record["resolved_merged_value"])
            else record["raw_value"]
        )
        for record in raw_cells
    }
    warnings: list[ExtractionWarning] = []
    for label in expected_labels:
        if normalize_label(label) not in available:
            warnings.append(
                ExtractionWarning(
                    warning_code="SH001",
                    severity=WarningSeverity.ERROR,
                    message=f"Expected anchor label is missing: {label!r}.",
                    workbook_id=workbook_id,
                    sheet_name=worksheet.title,
                    suggested_action=(
                        "Confirm the template version or update the sheet rule."
                    ),
                )
            )
    return warnings


def group_raw_cells_by_sheet(
    records: list[dict[str, Any]],
) -> dict[str, list[dict[str, Any]]]:
    """Group raw-cell records by worksheet name."""

    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        grouped[record["sheet_name"]].append(record)
    return dict(grouped)
