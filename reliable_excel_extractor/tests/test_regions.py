"""Region-detection tests."""

from pathlib import Path

from openpyxl import Workbook

from excel_extractor.config import ExtractionConfig
from excel_extractor.regions import detect_candidate_blocks
from excel_extractor.workbook import extract_raw_cells, load_workbook_pair


def test_side_by_side_tables_are_split(tmp_path: Path) -> None:
    path = tmp_path / "side_by_side.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"

    left = [
        ["KPI", "Value"],
        ["Revenue", 10],
        ["Cost", 7],
    ]
    right = [
        ["Metric", "Amount"],
        ["Headcount", 20],
        ["Sites", 4],
    ]
    for row_index, row in enumerate(left, start=1):
        for column_index, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=column_index, value=value)
    for row_index, row in enumerate(right, start=1):
        for column_index, value in enumerate(row, start=4):
            sheet.cell(row=row_index, column=column_index, value=value)
    workbook.save(path)

    formula_workbook, value_workbook = load_workbook_pair(path)
    raw_cells, _ = extract_raw_cells(
        formula_workbook["Data"],
        value_workbook["Data"],
        "test",
        ExtractionConfig(),
    )
    candidates = detect_candidate_blocks(
        formula_workbook["Data"],
        raw_cells,
        ExtractionConfig(),
    )

    assert [candidate.source_range for candidate in candidates] == [
        "A1:B3",
        "D1:E3",
    ]
