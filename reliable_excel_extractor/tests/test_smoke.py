"""Smoke tests for the extraction pipeline."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from excel_extractor import ExcelExtractor, ExtractionConfig


def _build_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "KPI Summary"

    sheet.merge_cells("B2:F2")
    sheet["B2"] = "KPI Performance"
    sheet["B2"].font = Font(bold=True, size=14)

    sheet.merge_cells("D3:E3")
    sheet["D3"] = 2025
    sheet["F3"] = 2026
    sheet["B4"] = "Region"
    sheet["C4"] = "KPI"
    sheet["D4"] = "Q1 2025"
    sheet["E4"] = "Q2 2025"
    sheet["F4"] = "Q1 2026"
    for cell in sheet[4][1:6]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")

    rows = [
        ("France", "Revenue", 100, 110, 120),
        ("France", "Cost", 70, 75, 80),
        ("France", "Margin", 30, 35, 40),
        ("Total", None, "=SUM(D5:D7)", "=SUM(E5:E7)", "=SUM(F5:F7)"),
    ]
    for row_index, values in enumerate(rows, start=5):
        for column_index, value in enumerate(values, start=2):
            sheet.cell(row=row_index, column=column_index, value=value)

    record = workbook.create_sheet("Record Data")
    record.append(["Entity", "Period", "KPI", "Value", "Unit"])
    record.append(["France", "Q1 2025", "Revenue", 100, "EURm"])
    record.append(["France", "Q1 2025", "Cost", 70, "EURm"])
    table = Table(displayName="RecordTable", ref="A1:E3")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    record.add_table(table)

    notes = workbook.create_sheet("Notes")
    notes["A1"] = "Notes"
    notes["A2"] = "Values are reported in millions."

    workbook.save(path)


def test_smoke(tmp_path: Path) -> None:
    workbook_path = tmp_path / "complex.xlsx"
    _build_workbook(workbook_path)

    result = ExcelExtractor(
        ExtractionConfig(accept_confidence=0.65)
    ).extract(workbook_path)

    assert len(result["sheet_inventory"]) == 3
    assert not result["raw_cells"].empty
    assert not result["detected_blocks"].empty
    observations = result["analytical_observations"]
    blocks = result["detected_blocks"]
    warnings = result["extraction_warnings"]

    assert not observations.empty
    assert {
        "source_sheet",
        "source_value_cell",
        "source_range",
        "extraction_confidence",
    }.issubset(observations.columns)

    kpi_block = blocks.loc[blocks["sheet_name"] == "KPI Summary"].iloc[0]
    assert kpi_block["block_class"] == "cross_tab_matrix"
    assert kpi_block["title_row_count"] == 1
    assert kpi_block["header_depth"] == 2

    kpi_observations = observations.loc[
        observations["source_sheet"] == "KPI Summary"
    ]
    assert len(kpi_observations) == 9
    assert set(kpi_observations["period_type"]) == {"quarter"}
    assert set(kpi_observations["period_year"]) == {2025, 2026}
    assert not kpi_observations["value"].map(lambda value: isinstance(value, str)).any()

    assert (warnings["warning_code"] == "CL001").sum() == 3
    assert not (warnings["warning_code"] == "VA002").any()
